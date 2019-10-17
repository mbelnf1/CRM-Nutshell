# -*- coding: utf-8 -*-
"""
Created on Tue Aug 13 14:28:37 2019

@author: mosta
"""
import csv
#import xlsxwriter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors, Protection, Color
#import openpyxl
#import xlrd
#from xlrd import open_workbook
import pandas as pd
import pandas
from openpyxl import Workbook

from openpyxl.cell import Cell
from openpyxl import load_workbook
from openpyxl.cell.cell import WriteOnlyCell
#import numpy as np
#from io import BytesIO as IO

with open(r'C:\Users\mosta\OneDrive\Documents\Programs\Python\Tuutkia\Nutshell CRM\tuutkia_crm.csv', 'r') as csvfile:
    spamreader = csv.reader(csvfile)
    newlist= list(spamreader)
    #print (newlist)
    
    


#Headers= ['Lastname','First name','Company','Title','Willing To share', 'Willing to introduce', 'Work phone', 'Work email', 'Work street', 'Work City', ' Work State', 'Work Zip', 'Personal Street', 'Personal City', 'Personal State', 'Personal Zip', 'Mobile Phone', 'Personal email', 'Note', 'Note Category']
#    
##Splitting an item in the list and adding it to a new list  
#namelist = [] # new, empty list
#for i in range(1, len(newlist)):
#    names = newlist[i][1][:].split() # this yields [first_name, last_name]
#    namelist.append([names[1], names[0]]) # [last_name, first_name]
#companylist=[]
#for i in range(1, len(newlist)):
#    p = re.compile(r'(.+)@(.+)\.(.+)')
#    test_str = newlist[i][3]
#    company= re.findall(p, test_str)
#    companyname= list(company[0][1])
#    companynom=''.join(companyname)
#    companylist.append(companynom)
#    # strip non-numeric characters'
#    workphone = []
#    wrkstreetaddress = []
#    workcityaddress = []
#    wrkstate = []
#    wrkzip = []
#    wrkmail = []
#    
#for i in range(1, len(newlist)):
#    phone = re.sub(r'\D', '', newlist[i][4])
#    # remove leading 1 (area codes never start with 1)
#    phone = phone.lstrip('1')
#    workingphone= '{}.{}.{}'.format(phone[0:3], phone[3:6], phone[6:])
#    workphone.append(workingphone)
#    wrkstraddress= newlist[i][10]
#    wrkstreetaddress.append(wrkstraddress)
#    wrkcityaddress= newlist[i][13]
#    workcityaddress.append(wrkcityaddress)
#    workstate= newlist[i][14]
#    wrkstate.append(workstate)
#    workzip=newlist[i][15]
#    wrkzip.append(workzip)
#    workemail=newlist[i][3]
#    wrkmail.append(workemail)
#
#    title = list()
#    Willingtoshare = list()
#    Willingtointroduce = list()
#    PersonalStreet = list()
#    PersonalCity = list()
#    PersonalState = list()
#    PersonalZip = list()
#    MobilePhones= list()
#    PersonalEmail = list()
#    Note = list()
#    NoteCategory = list()
#    
#Contact1= list()
#Contact1= [namelist[0][0]] + [namelist[1][0]] + [companylist[0]] + [title] + [Willingtoshare] + [Willingtointroduce] + [workphone] + [wrkmail] + []
#print(Contact1)
#Contact2= [namelist[1][0]]
#print(Contact2)
#Contact3= []
#
#
#    
#        
#        
#        
#        
#        
##namelist names
##companylist list of companies
##workphone list of formatted workphones
##teststr list of work emails
##straddress list of street addresses
#wb= Workbook()
#ws= wb.active
#ws.merge_cells('A1:D1')
#ws.merge_cells('G1:L1')
#ws.merge_cells('M1:R1')
#top_left_cell = ws['A1:D1']
#top_left_cell.value = "Contact Personal Information"
#top_center_cell = ws['G1:L1']
#top_center_cell.value = "Work Contact Information"
#top_right_cell = ws['M1:R1']
#top_right_cell.value = "Personal Contact Information"
#thin = Side(border_style="thin", color="000000")
#double = Side(border_style="double", color="ff0000")
#A2 = ws['A2'] 
#A2.value = "Lastname"
#A2.border = Border(top=double, left=double, right=double, bottom=double)
#B2 = ws['B2'] 
#B2.value = "Firstname"
#C2 = ws['C2'] 
#C2.value = "Company"
#C2.border = Border(top=double, left=double, right=double, bottom=double)
#D2 = ws['D2'] 
#D2.value = "Title"
#D2.border = Border(top=double, left=double, right=double, bottom=double)
#E2 = ws['E2'] 
#E2.value = "Willing to share"
#E2.border = Border(top=double, left=double, right=double, bottom=double)
#F2 = ws['F2'] 
#F2.value = "Willing to introduce"
#F2.border = Border(top=double, left=double, right=double, bottom=double)
#G2 = ws['G2'] 
#G2.value = "Work phone"
#G2.border = Border(top=double, left=double, right=double, bottom=double)
#H2 = ws['H2'] 
#H2.value = "Work email"
#H2.border = Border(top=double, left=double, right=double, bottom=double)
#I2 = ws['I2'] 
#I2.value = "Work Street"
#I2.border = Border(top=double, left=double, right=double, bottom=double)
#J2 = ws['J2'] 
#J2.value = "Work City"
#J2.border = Border(top=double, left=double, right=double, bottom=double)
#K2 = ws['K2'] 
#K2.value = "Work State"
#K2.border = Border(top=double, left=double, right=double, bottom=double)
#L2 = ws['L2'] 
#L2.value = "Work Zip"
#L2.border = Border(top=double, left=double, right=double, bottom=double)
#M2 = ws['M2'] 
#M2.value = "Personal Street"
#M2.border = Border(top=double, left=double, right=double, bottom=double)
#N2 = ws['N2'] 
#N2.value = "Personal City"
#N2.border = Border(top=double, left=double, right=double, bottom=double)
#O2 = ws['O2'] 
#O2.value = "Personal State"
#O2.border = Border(top=double, left=double, right=double, bottom=double)
#P2 = ws['P2'] 
#P2.value = "Personal Zip"
#P2.border = Border(top=double, left=double, right=double, bottom=double)
#Q2 = ws['Q2'] 
#Q2.value = "Mobile phone"
#Q2.border = Border(top=double, left=double, right=double, bottom=double)
#R2 = ws['R2'] 
#R2.value = "Personal email"
#R2.border = Border(top=double, left=double, right=double, bottom=double)
#S2 = ws['S2'] 
#S2.value = "Note"
#S2.border = Border(top=double, left=double, right=double, bottom=double)
#T2 = ws['T2'] 
#T2.value = "Note Category"
#T2.border = Border(top=double, left=double, right=double, bottom=double)

#workbook= xlsxwriter.Workbook('Test1.xlsx')
#worksheet= workbook.add_worksheet()
#
#
#worksheet.write('A2', 'LastName')
#worksheet.write('B2', 'FirstName')
#worksheet.write('C2', 'Company')
#worksheet.write('D2', 'Title')
#worksheet.write('E2', 'Willing To Share')
#worksheet.write('F2', 'Willing to introduce')
#worksheet.write('G2', 'Work phone')
#worksheet.write('H2', 'Work email')
#worksheet.write('I2', 'Work street')
#worksheet.write('J2', 'Work City')
#worksheet.write('K2', 'Work State')
#worksheet.write('L2', 'Work Zip')
#worksheet.write('M2', 'Personal Street')
#worksheet.write('N2', 'Personal City')
#worksheet.write('O2', 'Personal State')
#worksheet.write('P2', 'Personal Zip')
#worksheet.write('Q2', 'Mobile Phone')
#worksheet.write('R2', 'Personal email')
#worksheet.write('S2', 'Note')
#worksheet.write('T2', 'Note Category')
#
#cell_format = workbook.add_format()
#cell_format.set_border()   
#worksheet.set_column('A:T', 2, cell_format)
#cell_format3 = workbook.add_format()
#cell_format3.set_border()  
#worksheet.set_column('A:T', 3, cell_format3)
#cell_format4 = workbook.add_format()
#cell_format4.set_border()   
#worksheet.set_column('A:T', 4, cell_format4)   
#
#cell_format2 = workbook.add_format()
#cell_format2.set_border()
#cell_format2.set_align('bottom') 
#cell_format2.set_text_wrap()
#cell_format2.set_bg_color('f2f2f2')
##worksheet.set_row(2, cell_format2)
#worksheet.set_column('A:T', 2, cell_format2)
#worksheet.set_row(1, 80)
#worksheet.set_row(0, 26)
#worksheet.set_row(3, 40)
#worksheet.set_row(4, 60)
#worksheet.set_row(2, 40)
#worksheet.set_column('A:D', 11.29)
#worksheet.set_column('E:E', 9)
#worksheet.set_column('F:F', 8.14)
#worksheet.set_column('G:J', 11.29)
#worksheet.set_column('K:K', 5.29)
#worksheet.set_column('M:R', 11.29)
#worksheet.set_column('L:L', 8.29)
#worksheet.set_column('A:D', 11.29)
#worksheet.set_column('S:S', 36)
#worksheet.set_column('T:T', 8.14)
#
#
#
## Create a format to use in the merged range.
#merge_format = workbook.add_format({
#    'align': 'center',
#    'valign': 'bottom',
#    'fg_color': 'ffe699'})
#
#
## Merge 3 cells.
#worksheet.merge_range('A1:D1', 'Contact Personal Information', merge_format)
#
#merge_format2 = workbook.add_format({
#        'align': 'center',
#        'valign': 'bottom',
#        'fg_color': 'c6e0b4'})
#    
#worksheet.merge_range('G1:L1', 'Work Contact Information', merge_format2)
#
#merge_format3 = workbook.add_format({
#        'align': 'center',
#        'valign': 'bottom',
#        'fg_color': 'b4c6e7'})
#    
#worksheet.merge_range('M1:R1', 'Personal Contact Information', merge_format3)
#workbook2= xlsxwriter.Workbook('Test2.xlsx')
#worksheet2= workbook2.add_worksheet()
#
#
#worksheet2.write('A2', 'LastName')
#worksheet2.write('B2', 'FirstName')
#worksheet2.write('C2', 'Company')
#worksheet2.write('D2', 'Title')
#worksheet2.write('E2', 'Willing To Share')
#worksheet2.write('F2', 'Willing to introduce')
#worksheet2.write('G2', 'Work phone')
#worksheet2.write('H2', 'Work email')
#worksheet2.write('I2', 'Work street')
#worksheet2.write('J2', 'Work City')
#worksheet2.write('K2', 'Work State')
#worksheet2.write('L2', 'Work Zip')
#worksheet2.write('M2', 'Personal Street')
#worksheet2.write('N2', 'Personal City')
#worksheet2.write('O2', 'Personal State')
#worksheet2.write('P2', 'Personal Zip')
#worksheet2.write('Q2', 'Mobile Phone')
#worksheet2.write('R2', 'Personal email')
#worksheet2.write('S2', 'Note')
#worksheet2.write('T2', 'Note Category')
#
#cell_format9 = workbook2.add_format()
#cell_format9.set_border()   
#worksheet2.set_column('A:T', 2, cell_format9)
#cell_format5 = workbook2.add_format()
#cell_format5.set_border()  
#worksheet2.set_column('A:T', 3, cell_format5)
#cell_format7 = workbook2.add_format()
#cell_format7.set_border()   
#worksheet2.set_column('A:T', 4, cell_format7)   
#
#cell_format8 = workbook2.add_format()
#cell_format8.set_border()
#cell_format8.set_align('bottom') 
#cell_format8.set_text_wrap()
#cell_format8.set_bg_color('f2f2f2')
##worksheet.set_row(2, cell_format2)
#worksheet2.set_column('A:T', 2, cell_format8)
#worksheet2.set_row(1, 80)
#worksheet2.set_row(0, 26)
#worksheet2.set_row(3, 40)
#worksheet2.set_row(4, 60)
#worksheet2.set_row(2, 40)
#worksheet2.set_column('A:D', 11.29)
#worksheet2.set_column('E:E', 9)
#worksheet2.set_column('F:F', 8.14)
#worksheet2.set_column('G:J', 11.29)
#worksheet2.set_column('K:K', 5.29)
#worksheet2.set_column('M:R', 11.29)
#worksheet2.set_column('L:L', 8.29)
#worksheet2.set_column('A:D', 11.29)
#worksheet2.set_column('S:S', 36)
#worksheet2.set_column('T:T', 8.14)
#
#
#
## Create a format to use in the merged range.
#merge_format11 = workbook2.add_format({
#    'align': 'center',
#    'valign': 'bottom',
#    'fg_color': 'ffe699'})
#
#
## Merge 3 cells.
#worksheet2.merge_range('A1:D1', 'Contact Personal Information', merge_format11)
#
#merge_format12 = workbook2.add_format({
#        'align': 'center',
#        'valign': 'bottom',
#        'fg_color': 'c6e0b4'})
#    
#worksheet2.merge_range('G1:L1', 'Work Contact Information', merge_format12)
#
#merge_format13 = workbook2.add_format({
#        'align': 'center',
#        'valign': 'bottom',
#        'fg_color': 'b4c6e7'})
#    
#worksheet2.merge_range('M1:R1', 'Personal Contact Information', merge_format13)    


df = pd.DataFrame(newlist[1:], columns=newlist[0])
addresses = df.address_1.tolist()


df[['name', 'ï»¿id']] = df.name.str.split(' ', expand=True)

# rename id
df.rename(columns={'ï»¿id': 'Lastname','name': 'Firstname','lastContactedTime': 'Company','email':'Work_email', 'other_phones':'Personal_City', 'address_1':'Work_Street', 'address_2':'Personal_Zip', 'address_3':'Personal_Street', 'city':'Work_City', 'state':'Work_State', 'postal_code':'Work_Zip', 'tags':'Personal_email'}, inplace=True)
company2 = [' ', ' ', ' ',' ']
df['Company2'] = company2
df[['Company2','Company']] = df.Work_email.str.split('@', expand=True)

df.Company = [x.strip('.com') for x in df.Company]

del df['Company2']
df['Company'] = df.Company.str.capitalize()
del df['phone_phones']
#print(df)
 
note =  [' ', ' ', ' ',' ']
df['Note'] = note


notecat =  [' ', ' ', ' ',' ']
df['Note_Category'] = notecat

title =  [' ', ' ', ' ',' ']
df['Title'] = title

Willingtoshare =  [' ', ' ', ' ',' ']
df['Willing_to_share'] = Willingtoshare

Willingtointroduce =  [' ', ' ', ' ',' ']
df['Willing_to_introduce'] = Willingtointroduce

Personalstate =[' ', ' ', ' ',' ']
df['Personal_State'] = Personalstate


#df.Willing_to_share = [x.strip(' ') for x in df.Willing_to_share]
#df.Willing_to_introduce = [x.strip(' ') for x in df.Willing_to_introduce]




# split country_code from phone_phones
df[['country_code', 'work_phones']] = df.work_phones.str.split(' ', expand=True)
df[['country_code', 'mobile_phones']] = df.mobile_phones.str.split(' ', expand=True)
del df['country_code']
del df['country']
del df['fax_phones']
del df['home_phones']


#print (df.Personal_email.to_string(index=True))

#mobilephone =  [' ', ' ', ' ']
#df['Mobile_phone'] = mobilephone

print(df)
df = df[['Lastname', 'Firstname','Company','Title','Willing_to_share','Willing_to_introduce','work_phones','Work_email','Work_Street','Work_City','Work_State','Work_Zip','Personal_Street','Personal_City','Personal_State','Personal_Zip','mobile_phones','Personal_email','Note','Note_Category']]
#print(df)
#data=df.values.tolist()
#print(data2)
#print(data[0][0])
columns=df.columns
#print(columns)
#row = 2
#column = 0
#print (df.Willing_to_introduce.to_string(index=True)) 
#df.loc[(df.Lastname != ' ') & (df.Firstname != ' ') & ((df.Company != ' ') & (df.Title != ' ') | (df.Company = ' ')) & ((df.work_phones != ' ') | (df.mobile_phones != ' ') & (df.Personal_Street != ' ') & (df.Personal_City != ' ') & (df.Personal_State != ' ') & (df.Personal_Zip != ' ') | (df.Work_Street != ' ') & (df.Work_City != ' ') & (df.Work_State != ' ') & (df.Work_Zip != ' ')& (df.Work_email != ' ') | (df.Personal_email != ' ')),worksheet.write(row, column, df) & row += 1 | row += 1 & worksheet2.write(row, column, df)]:
#if (df['Lastname'] != ' ' & df['Firstname'] != ' ') & ((df['Company'] != ' ' & df['Title'] != ' ') | (df['Company'] == ' ') & (df['work_phones'] != ' ') | (df['mobile_phones'] != ' ') & (df['Personal_Street'] != ' ' & df['Personal_City'] != ' ' & df['Personal_State'] != ' ' & df['Personal_Zip'] != ' ') | (df['Work_Street'] != ' ' & df['Work_City'] != ' ' & df['Work_State'] != ' ' & df['Work_Zip'] != ' ' & (df['Work_email'] != ' ' | df['Personal_email'] != ' '))):
   #df.rename(columns={'Willing_to_share':'Willing to share','Willing_to_introduce':'Willing to introduce','work_phones': 'Work phone','Work_email':'Work email','Work_Street':'Work Street','Work_City':'Work City','Work_State':'Work State','Work_Zip':'Work Zip','Personal_Street':'Personal Street','Personal_City':'Personal  City','Personal_State':'Personal  State','Personal_Zip':'Personal  Zip','mobile_phones':'Mobile phone','Personal_email':'Personal email','Note_Category':'Note Category'}, inplace=True)    
## iterating through content list   
    ## write operation perform 
    ## worksheet.write(row, column, df) 
    ## incrementing the value of row by one 
    ## with each iteratons. 
    #row += 1
#else:
        
 #   worksheet2.write(row, column, df)
  #  row += 1
#sf=df.columns.tolist()
#book = load_workbook('Test1.xlsx')
#writer= pd.ExcelWriter('Test1.xlsx',engine='xlsxwriter') 
#writer.book = book
##writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#
#book2 = load_workbook('Test2.xlsx')
#writer2= pd.ExcelWriter('Test2.xlsx',engine='xlsxwriter')
#writer2.book2 = book2
##writer2.sheets = dict((ws.title, ws) for ws in book2.worksheets)
#
#workbook = writer.book
#workbook2 = writer2.book2
#
#worksheet = writer.sheets['Sheet1']
#worksheet2 = writer.sheets['Sheet2']
#
#value = data
for column in df:
   #for col, sf in enumerate(data):
   if (column[0] != ' ' and column[1] != ' ') and ((column[2] != ' ' and column[3] != ' ') or (column[2] == ' ') and (column[6] != ' ') or (column[16] != ' ') and (column[12] != ' ' and column[13] != ' ' and column[14] != ' ' and column[15] != ' ') or (column[8] != ' ' and column[9] != ' ' and column[10] != ' ' and column[11] != ' ' and (column[7] != ' ' or column[17] != ' '))):
       wb = Workbook()
       ws = wb.active
       wb2 = Workbook()
       ws2 = wb2.active
       ws.merge_cells('A1:D1')
       ws2.merge_cells('A1:D1')
       ws.merge_cells('G1:L1')
       ws2.merge_cells('G1:L1')
       ws.merge_cells('M1:R1')
       ws2.merge_cells('M1:R1')
       A1 = ws['A1']
       A12 = ws2['A1']
       A1.value = "Contact Personal Information"
       A12.value = "Contact Personal Information"
       A1.font = Font(size=11)
       A12.font = Font(size=11)
       yellowFill = PatternFill(start_color='ffe699',
                   end_color='ffe699',
                   fill_type='solid')
       A1.fill = yellowFill
       A12.fill = yellowFill
       A1.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       A12.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       #A1.alignment.vertical = "Bottom"
       #A1.alignment.horizontal = "Center"
       #A1.fill = PatternFill("solid", fgColor="ffe699")
       G1 = ws['G1']
       G12 = ws2['G1']
       G1.value = "Work Contact Information"
       G12.value = "Work Contact Information"
       G1.font = Font(size=11)
       G12.font = Font(size=11)
       greenFill = PatternFill(start_color='C6E0B4',
                   end_color='C6E0B4',
                   fill_type='solid')
       G1.fill = greenFill
       G12.fill = greenFill
       G1.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       G12.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       M1 = ws['M1']
       M12 = ws2['M1']
       M1.value = "Personal Contact Information"
       M12.value = "Personal Contact Information"
       M1.font = Font(size=11)
       M12.font = Font(size=11)
       blueFill = PatternFill(start_color='B4C6E7',
                   end_color='B4C6E7',
                   fill_type='solid')
       M1.fill = blueFill
       M12.fill = blueFill
       M1.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       M12.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       greyFill = PatternFill(start_color='f2f2f2',
                   end_color='f2f2f2',
                   fill_type='solid')
       
       ws.row_dimensions[1].height = 19.5
       ws2.row_dimensions[1].height = 19.5
       ws.row_dimensions[2].height = 60
       ws2.row_dimensions[2].height = 60
       ws.column_dimensions['A'].width = 11.29
       ws2.column_dimensions['A'].width = 11.29
       ws.column_dimensions['B'].width = 11.29
       ws2.column_dimensions['B'].width = 11.29
       ws.column_dimensions['C'].width = 11.29
       ws2.column_dimensions['C'].width = 11.29
       ws.column_dimensions['D'].width = 11.29
       ws2.column_dimensions['D'].width = 11.29
       ws.column_dimensions['E'].width = 9
       ws2.column_dimensions['E'].width = 9
       ws.column_dimensions['F'].width = 8.14
       ws2.column_dimensions['F'].width = 8.14
       ws.column_dimensions['G'].width = 11.29
       ws2.column_dimensions['G'].width = 11.29
       ws.column_dimensions['H'].width = 11.29
       ws2.column_dimensions['H'].width = 11.29
       ws.column_dimensions['I'].width = 11.29
       ws2.column_dimensions['I'].width = 11.29
       ws.column_dimensions['J'].width = 11.29
       ws2.column_dimensions['J'].width = 11.29
       ws.column_dimensions['K'].width = 5.29
       ws2.column_dimensions['K'].width = 5.29
       ws.column_dimensions['L'].width = 8.29
       ws2.column_dimensions['L'].width = 8.29
       ws.column_dimensions['M'].width = 11.29
       ws2.column_dimensions['M'].width = 11.29
       ws.column_dimensions['N'].width = 11.29
       ws2.column_dimensions['N'].width = 11.29
       ws.column_dimensions['O'].width = 11.29
       ws2.column_dimensions['O'].width = 11.29
       ws.column_dimensions['P'].width = 11.29
       ws2.column_dimensions['P'].width = 11.29
       ws.column_dimensions['Q'].width = 11.29
       ws2.column_dimensions['Q'].width = 11.29
       ws.column_dimensions['R'].width = 11.29
       ws2.column_dimensions['R'].width = 11.29
       ws.column_dimensions['S'].width = 36
       ws2.column_dimensions['S'].width = 36
       ws.column_dimensions['T'].width = 8.14
       ws2.column_dimensions['T'].width = 8.14
       thin = Side(border_style="thin", color="000000")
       double = Side(border_style="double", color="000000")
       A2 = ws['A2'] 
       A22 = ws2['A2'] 
       A2.value = "Lastname"
       A22.value = "Lastname"
       A2.font = Font(size=11)
       A22.font = Font(size=11)
       A2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A2.fill = greyFill
       A22.fill = greyFill
       A2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       A22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B2 = ws['B2'] 
       B22 = ws2['B2'] 
       B2.value = "Firstname"
       B22.value = "Firstname"
       B2.font = Font(size=11)
       B22.font = Font(size=11)
       B2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B2.fill = greyFill
       B22.fill = greyFill
       B2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C2 = ws['C2'] 
       C22 = ws2['C2']
       C2.value = "Company"
       C22.value = "Company"
       C2.font = Font(size=11)
       C22.font = Font(size=11)
       C2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C2.fill = greyFill
       C22.fill = greyFill
       C2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D2 = ws['D2'] 
       D22 = ws2['D2'] 
       D2.value = "Title"
       D22.value = "Title"
       D2.font = Font(size=11)
       D22.font = Font(size=11)
       D2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D2.fill = greyFill
       D22.fill = greyFill
       D2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E2 = ws['E2'] 
       E22 = ws2['E2'] 
       E2.value = "Willing to share"
       E22.value = "Willing to share"
       E2.font = Font(size=11)
       E22.font = Font(size=11)
       E2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E2.fill = greyFill
       E22.fill = greyFill
       E2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F2 = ws['F2'] 
       F22 = ws2['F2'] 
       F2.value = "Willing to introduce"
       F22.value = "Willing to introduce"
       F2.font = Font(size=11)
       F22.font = Font(size=11)
       F2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F2.fill = greyFill
       F22.fill = greyFill
       F2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G2 = ws['G2'] 
       G22 = ws2['G2'] 
       G2.value = "Work phone"
       G22.value = "Work phone"
       G2.font = Font(size=11)
       G22.font = Font(size=11)
       G2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G2.fill = greyFill
       G22.fill = greyFill
       G2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H2 = ws['H2'] 
       H22 = ws2['H2']
       H2.value = "Work email"
       H22.value = "Work email"
       H2.font = Font(size=11)
       H22.font = Font(size=11)
       H2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H2.fill = greyFill
       H22.fill = greyFill
       H2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I2 = ws['I2'] 
       I22 = ws2['I2']
       I2.value = "Work Street"
       I22.value = "Work Street"
       I2.font = Font(size=11)
       I22.font = Font(size=11)
       I2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I2.fill = greyFill
       I22.fill = greyFill
       I2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J2 = ws['J2'] 
       J22 = ws2['J2'] 
       J2.value = "Work City"
       J22.value = "Work City"
       J2.font = Font(size=11)
       J22.font = Font(size=11)
       J2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J2.fill = greyFill
       J22.fill = greyFill
       J2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K2 = ws['K2'] 
       K22 = ws2['K2'] 
       K2.value = "Work State"
       K22.value = "Work State"
       K2.font = Font(size=11)
       K22.font = Font(size=11)
       K2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K2.fill = greyFill
       K22.fill = greyFill
       K2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L2 = ws['L2'] 
       L22 = ws2['L2'] 
       L2.value = "Work Zip"
       L22.value = "Work Zip"
       L2.font = Font(size=11)
       L22.font = Font(size=11)
       L2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L2.fill = greyFill
       L22.fill = greyFill
       L2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M2 = ws['M2'] 
       M22 = ws2['M2'] 
       M2.value = "Personal Street"
       M22.value = "Personal Street"
       M2.font = Font(size=11)
       M22.font = Font(size=11)
       M2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M2.fill = greyFill
       M22.fill = greyFill
       M2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N2 = ws['N2'] 
       N22 = ws2['N2']
       N2.value = "Personal City"
       N22.value = "Personal City"
       N2.font = Font(size=11)
       N22.font = Font(size=11)
       N2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N2.fill = greyFill
       N22.fill = greyFill
       N2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O2 = ws['O2'] 
       O22 = ws2['O2']
       O2.value = "Personal State"
       O22.value = "Personal State"
       O2.font = Font(size=11)
       O22.font = Font(size=11)
       O2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O2.fill = greyFill
       O22.fill = greyFill
       O2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P2 = ws['P2'] 
       P22 = ws2['P2'] 
       P2.value = "Personal Zip"
       P22.value = "Personal Zip"
       P2.font = Font(size=11)
       P22.font = Font(size=11)
       P2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P2.fill = greyFill
       P22.fill = greyFill
       P2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q2 = ws['Q2'] 
       Q22 = ws2['Q2']
       Q2.value = "Mobile phone"
       Q22.value = "Mobile phone"
       Q2.font = Font(size=11)
       Q22.font = Font(size=11)
       Q2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q2.fill = greyFill
       Q22.fill = greyFill
       Q2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R2 = ws['R2'] 
       R22 = ws2['R2'] 
       R2.value = "Personal email"
       R22.value = "Personal email"
       R2.font = Font(size=11)
       R22.font = Font(size=11)
       R2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R2.fill = greyFill
       R22.fill = greyFill
       R2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S2 = ws['S2'] 
       S22 = ws2['S2'] 
       S2.value = "Note"
       S22.value = "Note"
       S2.font = Font(size=11)
       S22.font = Font(size=11)
       S2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S2.fill = greyFill
       S22.fill = greyFill
       S2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T2 = ws['T2'] 
       T22 = ws2['T2'] 
       T2.value = "Note Category"
       T22.value = "Note Category"
       T2.font = Font(size=11)
       T22.font = Font(size=11)
       T2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T2.fill = greyFill
       T22.fill = greyFill
       T2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       
     
       for r in dataframe_to_rows(df, index=False, header=False):
           
           ws.append(r)


       A3 = ws['A3'] 
       A3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       A4 = ws['A4'] 
       A4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       A5 = ws['A5'] 
       A5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       A6 = ws['A6'] 
       A6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B3 = ws['B3'] 
       B3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B4 = ws['B4'] 
       B4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B5 = ws['B5'] 
       B5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B6 = ws['B6'] 
       B6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C3 = ws['C3'] 
       C3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C4 = ws['C4'] 
       C4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C5 = ws['C5'] 
       C5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C6 = ws['C6'] 
       C6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D3 = ws['D3'] 
       D3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D4 = ws['D4'] 
       D4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D5 = ws['D5'] 
       D5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D6 = ws['D6'] 
       D6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E3 = ws['E3'] 
       E3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E4 = ws['E4'] 
       E4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E5 = ws['E5'] 
       E5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E6 = ws['E6'] 
       E6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F3 = ws['F3'] 
       F3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F4 = ws['F4'] 
       F4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F5 = ws['F5'] 
       F5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F6 = ws['F6'] 
       F6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G3 = ws['G3'] 
       G3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G4 = ws['G4'] 
       G4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G5 = ws['G5'] 
       G5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G6 = ws['G6'] 
       G6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H3 = ws['H3'] 
       H3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H4 = ws['H4'] 
       H4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H5 = ws['H5'] 
       H5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H6 = ws['H6'] 
       H6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I3 = ws['I3'] 
       I3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I4 = ws['I4'] 
       I4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I5 = ws['I5'] 
       I5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I6 = ws['I6'] 
       I6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J3 = ws['J3'] 
       J3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J4 = ws['J4'] 
       J4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J5 = ws['J5'] 
       J5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J6 = ws['J6'] 
       J6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K3 = ws['K3'] 
       K3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K4 = ws['K4'] 
       K4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K5 = ws['K5'] 
       K5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K6 = ws['K6'] 
       K6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L3 = ws['L3'] 
       L3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L4 = ws['L4'] 
       L4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L5 = ws['L5'] 
       L5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L6 = ws['L6'] 
       L6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M3 = ws['M3'] 
       M3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M4 = ws['M4'] 
       M4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M5 = ws['M5'] 
       M5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M6 = ws['M6'] 
       M6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N3 = ws['N3'] 
       N3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N4 = ws['N4'] 
       N4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N5 = ws['N5'] 
       N5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N6 = ws['N6'] 
       N6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O3 = ws['O3'] 
       O3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O4 = ws['O4'] 
       O4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O5 = ws['O5'] 
       O5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O6 = ws['O6'] 
       O6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P3 = ws['P3'] 
       P3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P4 = ws['P4'] 
       P4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P5 = ws['P5'] 
       P5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P6 = ws['P6'] 
       P6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q3 = ws['Q3'] 
       Q3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q4 = ws['Q4'] 
       Q4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q5 = ws['Q5'] 
       Q5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q6 = ws['Q6'] 
       Q6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R3 = ws['R3'] 
       R3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R4 = ws['R4'] 
       R4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R5 = ws['R5'] 
       R5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R6 = ws['R6'] 
       R6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S3 = ws['S3'] 
       S3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S4 = ws['S4'] 
       S4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S5 = ws['S5'] 
       S5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S6 = ws['S6'] 
       S6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T3 = ws['T3'] 
       T3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T4 = ws['T4'] 
       T4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T5 = ws['T5'] 
       T5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T6 = ws['T6'] 
       T6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       ws.row_dimensions[3].height = 30
       ws.row_dimensions[4].height = 30
       ws.row_dimensions[5].height = 30
       ws.row_dimensions[6].height = 45
           
           
       
               
               
          

       #for cell in ws['A'] + ws[1]:
          #cell.style = 'Pandas'

       wb.save("Accepted Contacts.xlsx")
       wb2.save("Rejected Contacts.xlsx")
       
       

#       cell = WriteOnlyCell(ws)
#       cell.style = 'Pandas'
#       cell2 = WriteOnlyCell(ws2)
#       cell2.style = 'Pandas'
#
#       def format_third_row(row, cell):
#
#            for c in row:
#                cell.value = c
#                yield cell
#
#       rows = dataframe_to_rows(df, index=False, header=False)
#       third_row = format_third_row(next(rows), cell)
#       ws.append(third_row)
#
#       for row in rows:
#           row = list(row)
#           cell.value = row[3]
#           row[3] = cell
#           ws.append(row)
#
#       wb.save("Test1.xlsx")
#       wb2.save("Test2.xlsx")
#       #worksheet.write('A3', data)
#                #df.to_excel(writer, header = False, startcol=1, startr
   else:
       wb = Workbook()
       ws = wb.active
       wb2 = Workbook()
       ws2 = wb2.active
       ws.merge_cells('A1:D1')
       ws2.merge_cells('A1:D1')
       ws.merge_cells('G1:L1')
       ws2.merge_cells('G1:L1')
       ws.merge_cells('M1:R1')
       ws2.merge_cells('M1:R1')
       A1 = ws['A1']
       A12 = ws2['A1']
       A1.value = "Contact Personal Information"
       A12.value = "Contact Personal Information"
       A1.font = Font(size=11)
       A12.font = Font(size=11)
       yellowFill = PatternFill(start_color='ffe699',
                   end_color='ffe699',
                   fill_type='solid')
       A1.fill = yellowFill
       A12.fill = yellowFill
       A1.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       A12.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       #A1.alignment.vertical = "Bottom"
       #A1.alignment.horizontal = "Center"
       #A1.fill = PatternFill("solid", fgColor="ffe699")
       G1 = ws['G1']
       G12 = ws2['G1']
       G1.value = "Work Contact Information"
       G12.value = "Work Contact Information"
       G1.font = Font(size=11)
       G12.font = Font(size=11)
       greenFill = PatternFill(start_color='C6E0B4',
                   end_color='C6E0B4',
                   fill_type='solid')
       G1.fill = greenFill
       G12.fill = greenFill
       G1.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       G12.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       M1 = ws['M1']
       M12 = ws2['M1']
       M1.value = "Personal Contact Information"
       M12.value = "Personal Contact Information"
       M1.font = Font(size=11)
       M12.font = Font(size=11)
       blueFill = PatternFill(start_color='B4C6E7',
                   end_color='B4C6E7',
                   fill_type='solid')
       M1.fill = blueFill
       M12.fill = blueFill
       M1.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       M12.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
       greyFill = PatternFill(start_color='f2f2f2',
                   end_color='f2f2f2',
                   fill_type='solid')
       
       ws.row_dimensions[1].height = 19.5
       ws2.row_dimensions[1].height = 19.5
       ws.row_dimensions[2].height = 60
       ws2.row_dimensions[2].height = 60
       ws.column_dimensions['A'].width = 11.29
       ws2.column_dimensions['A'].width = 11.29
       ws.column_dimensions['B'].width = 11.29
       ws2.column_dimensions['B'].width = 11.29
       ws.column_dimensions['C'].width = 11.29
       ws2.column_dimensions['C'].width = 11.29
       ws.column_dimensions['D'].width = 11.29
       ws2.column_dimensions['D'].width = 11.29
       ws.column_dimensions['E'].width = 9
       ws2.column_dimensions['E'].width = 9
       ws.column_dimensions['F'].width = 8.14
       ws2.column_dimensions['F'].width = 8.14
       ws.column_dimensions['G'].width = 11.29
       ws2.column_dimensions['G'].width = 11.29
       ws.column_dimensions['H'].width = 11.29
       ws2.column_dimensions['H'].width = 11.29
       ws.column_dimensions['I'].width = 11.29
       ws2.column_dimensions['I'].width = 11.29
       ws.column_dimensions['J'].width = 11.29
       ws2.column_dimensions['J'].width = 11.29
       ws.column_dimensions['K'].width = 5.29
       ws2.column_dimensions['K'].width = 5.29
       ws.column_dimensions['L'].width = 8.29
       ws2.column_dimensions['L'].width = 8.29
       ws.column_dimensions['M'].width = 11.29
       ws2.column_dimensions['M'].width = 11.29
       ws.column_dimensions['N'].width = 11.29
       ws2.column_dimensions['N'].width = 11.29
       ws.column_dimensions['O'].width = 11.29
       ws2.column_dimensions['O'].width = 11.29
       ws.column_dimensions['P'].width = 11.29
       ws2.column_dimensions['P'].width = 11.29
       ws.column_dimensions['Q'].width = 11.29
       ws2.column_dimensions['Q'].width = 11.29
       ws.column_dimensions['R'].width = 11.29
       ws2.column_dimensions['R'].width = 11.29
       ws.column_dimensions['S'].width = 36
       ws2.column_dimensions['S'].width = 36
       ws.column_dimensions['T'].width = 8.14
       ws2.column_dimensions['T'].width = 8.14
       thin = Side(border_style="thin", color="000000")
       double = Side(border_style="double", color="000000")
       A2 = ws['A2'] 
       A22 = ws2['A2'] 
       A2.value = "Lastname"
       A22.value = "Lastname"
       A2.font = Font(size=11)
       A22.font = Font(size=11)
       A2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A2.fill = greyFill
       A22.fill = greyFill
       A2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       A22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B2 = ws['B2'] 
       B22 = ws2['B2'] 
       B2.value = "Firstname"
       B22.value = "Firstname"
       B2.font = Font(size=11)
       B22.font = Font(size=11)
       B2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B2.fill = greyFill
       B22.fill = greyFill
       B2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C2 = ws['C2'] 
       C22 = ws2['C2']
       C2.value = "Company"
       C22.value = "Company"
       C2.font = Font(size=11)
       C22.font = Font(size=11)
       C2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C2.fill = greyFill
       C22.fill = greyFill
       C2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D2 = ws['D2'] 
       D22 = ws2['D2'] 
       D2.value = "Title"
       D22.value = "Title"
       D2.font = Font(size=11)
       D22.font = Font(size=11)
       D2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D2.fill = greyFill
       D22.fill = greyFill
       D2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E2 = ws['E2'] 
       E22 = ws2['E2'] 
       E2.value = "Willing to share"
       E22.value = "Willing to share"
       E2.font = Font(size=11)
       E22.font = Font(size=11)
       E2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E2.fill = greyFill
       E22.fill = greyFill
       E2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F2 = ws['F2'] 
       F22 = ws2['F2'] 
       F2.value = "Willing to introduce"
       F22.value = "Willing to introduce"
       F2.font = Font(size=11)
       F22.font = Font(size=11)
       F2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F2.fill = greyFill
       F22.fill = greyFill
       F2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G2 = ws['G2'] 
       G22 = ws2['G2'] 
       G2.value = "Work phone"
       G22.value = "Work phone"
       G2.font = Font(size=11)
       G22.font = Font(size=11)
       G2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G2.fill = greyFill
       G22.fill = greyFill
       G2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H2 = ws['H2'] 
       H22 = ws2['H2']
       H2.value = "Work email"
       H22.value = "Work email"
       H2.font = Font(size=11)
       H22.font = Font(size=11)
       H2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H2.fill = greyFill
       H22.fill = greyFill
       H2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I2 = ws['I2'] 
       I22 = ws2['I2']
       I2.value = "Work Street"
       I22.value = "Work Street"
       I2.font = Font(size=11)
       I22.font = Font(size=11)
       I2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I2.fill = greyFill
       I22.fill = greyFill
       I2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J2 = ws['J2'] 
       J22 = ws2['J2'] 
       J2.value = "Work City"
       J22.value = "Work City"
       J2.font = Font(size=11)
       J22.font = Font(size=11)
       J2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J2.fill = greyFill
       J22.fill = greyFill
       J2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K2 = ws['K2'] 
       K22 = ws2['K2'] 
       K2.value = "Work State"
       K22.value = "Work State"
       K2.font = Font(size=11)
       K22.font = Font(size=11)
       K2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K2.fill = greyFill
       K22.fill = greyFill
       K2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L2 = ws['L2'] 
       L22 = ws2['L2'] 
       L2.value = "Work Zip"
       L22.value = "Work Zip"
       L2.font = Font(size=11)
       L22.font = Font(size=11)
       L2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L2.fill = greyFill
       L22.fill = greyFill
       L2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M2 = ws['M2'] 
       M22 = ws2['M2'] 
       M2.value = "Personal Street"
       M22.value = "Personal Street"
       M2.font = Font(size=11)
       M22.font = Font(size=11)
       M2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M2.fill = greyFill
       M22.fill = greyFill
       M2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N2 = ws['N2'] 
       N22 = ws2['N2']
       N2.value = "Personal City"
       N22.value = "Personal City"
       N2.font = Font(size=11)
       N22.font = Font(size=11)
       N2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N2.fill = greyFill
       N22.fill = greyFill
       N2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O2 = ws['O2'] 
       O22 = ws2['O2']
       O2.value = "Personal State"
       O22.value = "Personal State"
       O2.font = Font(size=11)
       O22.font = Font(size=11)
       O2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O2.fill = greyFill
       O22.fill = greyFill
       O2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P2 = ws['P2'] 
       P22 = ws2['P2'] 
       P2.value = "Personal Zip"
       P22.value = "Personal Zip"
       P2.font = Font(size=11)
       P22.font = Font(size=11)
       P2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P2.fill = greyFill
       P22.fill = greyFill
       P2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q2 = ws['Q2'] 
       Q22 = ws2['Q2']
       Q2.value = "Mobile phone"
       Q22.value = "Mobile phone"
       Q2.font = Font(size=11)
       Q22.font = Font(size=11)
       Q2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q2.fill = greyFill
       Q22.fill = greyFill
       Q2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R2 = ws['R2'] 
       R22 = ws2['R2'] 
       R2.value = "Personal email"
       R22.value = "Personal email"
       R2.font = Font(size=11)
       R22.font = Font(size=11)
       R2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R2.fill = greyFill
       R22.fill = greyFill
       R2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S2 = ws['S2'] 
       S22 = ws2['S2'] 
       S2.value = "Note"
       S22.value = "Note"
       S2.font = Font(size=11)
       S22.font = Font(size=11)
       S2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S2.fill = greyFill
       S22.fill = greyFill
       S2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T2 = ws['T2'] 
       T22 = ws2['T2'] 
       T2.value = "Note Category"
       T22.value = "Note Category"
       T2.font = Font(size=11)
       T22.font = Font(size=11)
       T2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T2.fill = greyFill
       T22.fill = greyFill
       T2.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T22.alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       
     
       for r in dataframe_to_rows(df, index=False, header=False):
           
           ws.append(r)


       A3 = ws['A3'] 
       A3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       A4 = ws['A4'] 
       A4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       A5 = ws['A5'] 
       A5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       A6 = ws['A6'] 
       A6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       A6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B3 = ws['B3'] 
       B3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B4 = ws['B4'] 
       B4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B5 = ws['B5'] 
       B5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       B6 = ws['B6'] 
       B6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       B6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C3 = ws['C3'] 
       C3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C4 = ws['C4'] 
       C4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C5 = ws['C5'] 
       C5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       C6 = ws['C6'] 
       C6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       C6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D3 = ws['D3'] 
       D3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D4 = ws['D4'] 
       D4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D5 = ws['D5'] 
       D5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       D6 = ws['D6'] 
       D6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E3 = ws['E3'] 
       E3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E4 = ws['E4'] 
       E4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E5 = ws['E5'] 
       E5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       E6 = ws['E6'] 
       E6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       E6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F3 = ws['F3'] 
       F3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F4 = ws['F4'] 
       F4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F5 = ws['F5'] 
       F5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       F6 = ws['F6'] 
       F6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       F6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G3 = ws['G3'] 
       G3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G4 = ws['G4'] 
       G4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G5 = ws['G5'] 
       G5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       G6 = ws['G6'] 
       G6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       G6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H3 = ws['H3'] 
       H3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H4 = ws['H4'] 
       H4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H5 = ws['H5'] 
       H5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       H6 = ws['H6'] 
       H6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       H6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I3 = ws['I3'] 
       I3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I4 = ws['I4'] 
       I4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I5 = ws['I5'] 
       I5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       I6 = ws['I6'] 
       I6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       I6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J3 = ws['J3'] 
       J3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J4 = ws['J4'] 
       J4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J5 = ws['J5'] 
       J5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       J6 = ws['J6'] 
       J6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       J6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K3 = ws['K3'] 
       K3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K4 = ws['K4'] 
       K4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K5 = ws['K5'] 
       K5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       K6 = ws['K6'] 
       K6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       K6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L3 = ws['L3'] 
       L3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L4 = ws['L4'] 
       L4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L5 = ws['L5'] 
       L5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       L6 = ws['L6'] 
       L6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       L6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M3 = ws['M3'] 
       M3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M4 = ws['M4'] 
       M4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M5 = ws['M5'] 
       M5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       M6 = ws['M6'] 
       M6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       M6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N3 = ws['N3'] 
       N3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N4 = ws['N4'] 
       N4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N5 = ws['N5'] 
       N5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       N6 = ws['N6'] 
       N6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       N6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O3 = ws['O3'] 
       O3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O4 = ws['O4'] 
       O4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O5 = ws['O5'] 
       O5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       O6 = ws['O6'] 
       O6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       O6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P3 = ws['P3'] 
       P3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P4 = ws['P4'] 
       P4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P5 = ws['P5'] 
       P5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       D5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       P6 = ws['P6'] 
       P6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       P6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q3 = ws['Q3'] 
       Q3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q4 = ws['Q4'] 
       Q4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q5 = ws['Q5'] 
       Q5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       Q6 = ws['Q6'] 
       Q6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       Q6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R3 = ws['R3'] 
       R3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R4 = ws['R4'] 
       R4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R5 = ws['R5'] 
       R5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       R6 = ws['R6'] 
       R6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       R6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S3 = ws['S3'] 
       S3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S4 = ws['S4'] 
       S4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S5 = ws['S5'] 
       S5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       S6 = ws['S6'] 
       S6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       S6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T3 = ws['T3'] 
       T3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T3.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T4 = ws['T4'] 
       T4.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T4.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T5 = ws['T5'] 
       T5.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T5.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       T6 = ws['T6'] 
       T6.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       T6.alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
       ws.row_dimensions[3].height = 30
       ws.row_dimensions[4].height = 30
       ws.row_dimensions[5].height = 30
       ws.row_dimensions[6].height = 45
           
           
       
               
               
          

       #for cell in ws['A'] + ws[1]:
          #cell.style = 'Pandas'

       wb.save("Rejected Contacts.xlsx")
       wb2.save("Accepted Contacts.xlsx")
                #worksheet2.write('A3', data)
                #df.to_excel(writer2, header = False, startcol=1, startrow=3)


#print(data)
#row_num = 2
#col_num = 0
#col = 0
#sheet1 = worksheet.sheet_by_index(0)
#sheet2 = worksheet2.sheet_by_index(0)
#excel_file = IO()
#writer= pd.ExcelWriter('Test1.xlsx',engine='xlsxwriter')
#worksheet = writer.sheets['Sheet1']
#worksheet2 = writer.sheets['Sheet2'] 
#df.to_excel(writer, header = False, sheet_name= worksheet, startcol=1, startrow=3)  # startcol, startrow can be used in this condition
#df.to_excel(writer, header = False, sheet_name= worksheet2, startcol=1, startrow=3)

#worksheet = writer.sheets['Sheet1']
#worksheet2 = writer.sheets['Sheet1'] 
#for column in df:
#   for col, sf in enumerate(data):
 #      if (column[0] != ' ' and column[1] != ' ') and ((column[2] != ' ' and column[3] != ' ') or (column[2] == ' ') and (column[6] != ' ') or (column[16] != ' ') and (column[12] != ' ' and column[13] != ' ' and column[14] != ' ' and column[15] != ' ') or (column[8] != ' ' and column[9] != ' ' and column[10] != ' ' and column[11] != ' ' and (column[7] != ' ' or column[17] != ' '))):
  #        data.to_excel(writer, header = False, sheet_name='Sheet1', startcol=1, startrow=3)
   #    else:
    #      data.to_excel(writer, header = False, sheet_name='Sheet2', startcol=1, startrow=3)
#for column in df:
   #for col, sf in enumerate(data):
            #if (column[0] != ' ' and column[1] != ' ') and ((column[2] != ' ' and column[3] != ' ') or (column[2] == ' ') and (column[6] != ' ') or (column[16] != ' ') and (column[12] != ' ' and column[13] != ' ' and column[14] != ' ' and column[15] != ' ') or (column[8] != ' ' and column[9] != ' ' and column[10] != ' ' and column[11] != ' ' and (column[7] != ' ' or column[17] != ' '))):
            #writer = pd.ExcelWriter(worksheet, engine='xlsxwriter')
            #worksheet.write(2, 0, data) 
            #df.to_excel(writer, sheet_name = 'Sheet1', startrow = 2)
            #data.append(sheet1.cell(i,2).value)
            #worksheet.write_column(row_num, col_num, info)
                    #worksheet.write_row(2, 0, sf)
              #      worksheet.write("A3",df)
            
            #for row in data:
                #worksheet.append(row)
            #else:
            #writer = pd.ExcelWriter(worksheet2, engine = 'xlsxwriter')
            #worksheet2.write(2, 0, data) 
                    #worksheet2.write_row(2, 0, sf)
                    #worksheet2.write("A3",df) 
            #data.append(sheet2.cell(i,2).value)
            #worksheet2.write_column(row_num, col_num, info)
            #for row in data:
                #worksheet2.append(row)
                

#print (df.Willing to share.to_string(index=True))
#df.loc[(df.First_name == 'Bill') & (df.First_name == 'Emma'), 'name_match'] = 'Match' 




    

    
# 
#
#
#    
#
#wb.save("Test1.xlsx")
#wb2.save("Test2.xlsx")

#writer.save()
#writer.close()
#writer2.save()
#writer2.close()
#workbook.close()
#workbook2.close()

