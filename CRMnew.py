# -*- coding: utf-8 -*-
"""
Created on Fri Oct 18 13:29:23 2019

@author: mosta
"""

import csv
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors, Protection, Color
import pandas as pd
import pandas
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import numpy as np
from openpyxl import utils
from openpyxl.cell import Cell
from openpyxl import load_workbook
from openpyxl.cell.cell import WriteOnlyCell

with open(r'C:\Users\mosta\OneDrive\Documents\Programs\Python\Tuutkia\Nutshell CRM\tuutkia_crm.csv', 'r') as csvfile:
    spamreader = csv.reader(csvfile)
    newlist= list(spamreader)
    
df = pd.DataFrame(newlist[1:], columns=newlist[0])
addresses = df.address_1.tolist()


df[['name', 'ï»¿id']] = df.name.str.split(' ', expand=True)

# rename id
df.rename(columns={'ï»¿id': 'Lastname','name': 'Firstname','lastContactedTime': 'Company','email':'Work_email', 'other_phones':'Personal_City', 'address_1':'Work_Street', 'address_2':'Personal_Zip', 'address_3':'Personal_Street', 'city':'Work_City', 'state':'Work_State', 'postal_code':'Work_Zip', 'tags':'Personal_email'}, inplace=True)





del df['phone_phones']
#print(df)
 
df['Note'] = ''



df['Note_Category'] = ''


df['Title'] = ''


df['Willing_to_share'] = ''


df['Willing_to_introduce'] = ''


df['Personal_State'] = ''

df[['country_code', 'work_phones']] = df.work_phones.str.split(' ', expand=True)
df[['country_code', 'mobile_phones']] = df.mobile_phones.str.split(' ', expand=True)
del df['country_code']
del df['country']
del df['fax_phones']
del df['home_phones']


df = df[['Lastname', 'Firstname','Company','Title','Willing_to_share','Willing_to_introduce','work_phones','Work_email','Work_Street','Work_City','Work_State','Work_Zip','Personal_Street','Personal_City','Personal_State','Personal_Zip','mobile_phones','Personal_email','Note','Note_Category']]
df = df.applymap(lambda x: np.nan if isinstance(x, str) and x.strip() == '' else x)

wb = Workbook()
ws = wb.active
wb2 = Workbook()
ws2 = wb2.active
ws.merge_cells('A1:D1')
ws2.merge_cells('A1:D1')
ws.merge_cells('G1:L1')
ws2.merge_cells('G1:L1')
ws.merge_cells('M1:R1')
ws2.merge_cells('M1:R1')
A1 = ws['A1']
A12 = ws2['A1']
A1.value = "Contact Personal Information"
A12.value = "Contact Personal Information"
A1.font = Font(size=11)
A12.font = Font(size=11)
yellowFill = PatternFill(start_color='ffe699',
end_color='ffe699',
fill_type='solid')
A1.fill = yellowFill
A12.fill = yellowFill
A1.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=False,
 shrink_to_fit=False,
 indent=0)
A12.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=False,
 shrink_to_fit=False,
 indent=0)
G1 = ws['G1']
G12 = ws2['G1']
G1.value = "Work Contact Information"
G12.value = "Work Contact Information"
G1.font = Font(size=11)
G12.font = Font(size=11)
greenFill = PatternFill(start_color='C6E0B4',
   end_color='C6E0B4',
   fill_type='solid')
G1.fill = greenFill
G12.fill = greenFill
G1.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=False,
 shrink_to_fit=False,
 indent=0)
G12.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=False,
 shrink_to_fit=False,
 indent=0)
M1 = ws['M1']
M12 = ws2['M1']
M1.value = "Personal Contact Information"
M12.value = "Personal Contact Information"
M1.font = Font(size=11)
M12.font = Font(size=11)
blueFill = PatternFill(start_color='B4C6E7',
   end_color='B4C6E7',
   fill_type='solid')
M1.fill = blueFill
M12.fill = blueFill
M1.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=False,
 shrink_to_fit=False,
 indent=0)
M12.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=False,
 shrink_to_fit=False,
 indent=0)
greyFill = PatternFill(start_color='f2f2f2',
   end_color='f2f2f2',
   fill_type='solid')
   
ws.row_dimensions[1].height = 19.5
ws2.row_dimensions[1].height = 19.5
ws.row_dimensions[2].height = 60
ws2.row_dimensions[2].height = 60
ws.column_dimensions['A'].width = 11.29
ws2.column_dimensions['A'].width = 11.29
ws.column_dimensions['B'].width = 11.29
ws2.column_dimensions['B'].width = 11.29
ws.column_dimensions['C'].width = 11.29
ws2.column_dimensions['C'].width = 11.29
ws.column_dimensions['D'].width = 11.29
ws2.column_dimensions['D'].width = 11.29
ws.column_dimensions['E'].width = 9
ws2.column_dimensions['E'].width = 9
ws.column_dimensions['F'].width = 8.14
ws2.column_dimensions['F'].width = 8.14
ws.column_dimensions['G'].width = 11.29
ws2.column_dimensions['G'].width = 11.29
ws.column_dimensions['H'].width = 11.29
ws2.column_dimensions['H'].width = 11.29
ws.column_dimensions['I'].width = 11.29
ws2.column_dimensions['I'].width = 11.29
ws.column_dimensions['J'].width = 11.29
ws2.column_dimensions['J'].width = 11.29
ws.column_dimensions['K'].width = 5.29
ws2.column_dimensions['K'].width = 5.29
ws.column_dimensions['L'].width = 8.29
ws2.column_dimensions['L'].width = 8.29
ws.column_dimensions['M'].width = 11.29
ws2.column_dimensions['M'].width = 11.29
ws.column_dimensions['N'].width = 11.29
ws2.column_dimensions['N'].width = 11.29
ws.column_dimensions['O'].width = 11.29
ws2.column_dimensions['O'].width = 11.29
ws.column_dimensions['P'].width = 11.29
ws2.column_dimensions['P'].width = 11.29
ws.column_dimensions['Q'].width = 11.29
ws2.column_dimensions['Q'].width = 11.29
ws.column_dimensions['R'].width = 11.29
ws2.column_dimensions['R'].width = 11.29
ws.column_dimensions['S'].width = 36
ws2.column_dimensions['S'].width = 36
ws.column_dimensions['T'].width = 8.14
ws2.column_dimensions['T'].width = 8.14
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="000000")
A2 = ws['A2'] 
A22 = ws2['A2'] 
A2.value = "Lastname"
A22.value = "Lastname"
A2.font = Font(size=11)
A22.font = Font(size=11)
A2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
A22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
A2.fill = greyFill
A22.fill = greyFill
A2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
A22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
B2 = ws['B2'] 
B22 = ws2['B2'] 
B2.value = "Firstname"
B22.value = "Firstname"
B2.font = Font(size=11)
B22.font = Font(size=11)
B2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
B22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
B2.fill = greyFill
B22.fill = greyFill
B2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
B22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
C2 = ws['C2'] 
C22 = ws2['C2']
C2.value = "Company"
C22.value = "Company"
C2.font = Font(size=11)
C22.font = Font(size=11)
C2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
C22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
C2.fill = greyFill
C22.fill = greyFill
C2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
C22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
D2 = ws['D2'] 
D22 = ws2['D2'] 
D2.value = "Title"
D22.value = "Title"
D2.font = Font(size=11)
D22.font = Font(size=11)
D2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
D22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
D2.fill = greyFill
D22.fill = greyFill
D2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
D22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
E2 = ws['E2'] 
E22 = ws2['E2'] 
E2.value = "Willing to share"
E22.value = "Willing to share"
E2.font = Font(size=11)
E22.font = Font(size=11)
E2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
E22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
E2.fill = greyFill
E22.fill = greyFill
E2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
E22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
F2 = ws['F2'] 
F22 = ws2['F2'] 
F2.value = "Willing to introduce"
F22.value = "Willing to introduce"
F2.font = Font(size=11)
F22.font = Font(size=11)
F2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
F22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
F2.fill = greyFill
F22.fill = greyFill
F2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
F22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
G2 = ws['G2'] 
G22 = ws2['G2'] 
G2.value = "Work phone"
G22.value = "Work phone"
G2.font = Font(size=11)
G22.font = Font(size=11)
G2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
G22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
G2.fill = greyFill
G22.fill = greyFill
G2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
G22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
H2 = ws['H2'] 
H22 = ws2['H2']
H2.value = "Work email"
H22.value = "Work email"
H2.font = Font(size=11)
H22.font = Font(size=11)
H2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
H22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
H2.fill = greyFill
H22.fill = greyFill
H2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
H22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
I2 = ws['I2'] 
I22 = ws2['I2']
I2.value = "Work Street"
I22.value = "Work Street"
I2.font = Font(size=11)
I22.font = Font(size=11)
I2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
I22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
I2.fill = greyFill
I22.fill = greyFill
I2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
I22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
J2 = ws['J2'] 
J22 = ws2['J2'] 
J2.value = "Work City"
J22.value = "Work City"
J2.font = Font(size=11)
J22.font = Font(size=11)
J2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
J22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
J2.fill = greyFill
J22.fill = greyFill
J2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
J22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
K2 = ws['K2'] 
K22 = ws2['K2'] 
K2.value = "Work State"
K22.value = "Work State"
K2.font = Font(size=11)
K22.font = Font(size=11)
K2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
K22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
K2.fill = greyFill
K22.fill = greyFill
K2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
K22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
L2 = ws['L2'] 
L22 = ws2['L2'] 
L2.value = "Work Zip"
L22.value = "Work Zip"
L2.font = Font(size=11)
L22.font = Font(size=11)
L2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
L22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
L2.fill = greyFill
L22.fill = greyFill
L2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
L22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
M2 = ws['M2'] 
M22 = ws2['M2'] 
M2.value = "Personal Street"
M22.value = "Personal Street"
M2.font = Font(size=11)
M22.font = Font(size=11)
M2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
M22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
M2.fill = greyFill
M22.fill = greyFill
M2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
M22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
N2 = ws['N2'] 
N22 = ws2['N2']
N2.value = "Personal City"
N22.value = "Personal City"
N2.font = Font(size=11)
N22.font = Font(size=11)
N2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
N22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
N2.fill = greyFill
N22.fill = greyFill
N2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
N2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
O2 = ws['O2'] 
O22 = ws2['O2']
O2.value = "Personal State"
O22.value = "Personal State"
O2.font = Font(size=11)
O22.font = Font(size=11)
O2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
O22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
O2.fill = greyFill
O22.fill = greyFill
O2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
O22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
P2 = ws['P2'] 
P22 = ws2['P2'] 
P2.value = "Personal Zip"
P22.value = "Personal Zip"
P2.font = Font(size=11)
P22.font = Font(size=11)
P2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
P22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
P2.fill = greyFill
P22.fill = greyFill
P2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
P22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
Q2 = ws['Q2'] 
Q22 = ws2['Q2']
Q2.value = "Mobile phone"
Q22.value = "Mobile phone"
Q2.font = Font(size=11)
Q22.font = Font(size=11)
Q2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
Q22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
Q2.fill = greyFill
Q22.fill = greyFill
Q2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
Q22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
R2 = ws['R2'] 
R22 = ws2['R2'] 
R2.value = "Personal email"
R22.value = "Personal email"
R2.font = Font(size=11)
R22.font = Font(size=11)
R2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
R22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
R2.fill = greyFill
R22.fill = greyFill
R2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
R22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
S2 = ws['S2'] 
S22 = ws2['S2'] 
S2.value = "Note"
S22.value = "Note"
S2.font = Font(size=11)
S22.font = Font(size=11)
S2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
S22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
S2.fill = greyFill
S22.fill = greyFill
S2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
S22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
T2 = ws['T2'] 
T22 = ws2['T2'] 
T2.value = "Note Category"
T22.value = "Note Category"
T2.font = Font(size=11)
T22.font = Font(size=11)
T2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
T22.border = Border(top=thin, left=thin, right=thin, bottom=thin)
T2.fill = greyFill
T22.fill = greyFill
T2.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
T22.alignment=Alignment(horizontal='center',
 vertical='bottom',
 text_rotation=0,
 wrap_text=True,
 shrink_to_fit=False,
 indent=0)
   

r1 = df.dropna(subset=['Firstname', 'Lastname', ('work_phones' or 'mobile_phones') or (('Work_City','Work_Street','Work_State','Work_Zip') or ('Personal_Street','Personal_City','Personal_State','Personal_Zip')) or ('Work_email' or 'Personal_email')])

r2 = df.loc[(df['Firstname'].isnull()) | (df['Lastname'].isnull()) | (((df['work_phones'].isnull()) & (df['mobile_phones'].isnull())) | (((df['Work_Street'].isnull()) | (df['Work_City'].isnull()) | (df['Work_State'].isnull()) & (df['Work_Zip'].isnull())) | (df['Personal_Street'].isnull()) | (df['Personal_City'].isnull()) | (df['Personal_State'].isnull()) | (df['Personal_Zip'].isnull())) & (df['Work_email'].isnull()) & (df['Personal_email'].isnull()))]

for i in range(r1.shape[0]+3):
    ws.row_dimensions[i].height = 30

for r in dataframe_to_rows(r1, index=False, header=False):
   ws.append(r)
   

def set_border(ws, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows(cell_range)
    for row in rows:
        for cell in row:
            cell.border = border

set_border(ws, r1.shape[0])  

for i in range(r2.shape[0]+3):
    ws2.row_dimensions[i].height = 30

# =============================================================================
for r in dataframe_to_rows(r2, index=False, header=False):
    ws2.append(r)
  
def set_border(ws2, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws2.iter_rows(cell_range)
    for row in rows:
        for cell in row:
            cell.border = border

set_border(ws2, r2.shape[0]-2)   

wb.save("Accepted Contacts.xlsx")
file_loc = r'C:\Users\mosta\.spyder-py3\Accepted Contacts.xlsx'

ss = openpyxl.load_workbook(file_loc)
ss_sheet = ss.get_sheet_by_name('Sheet')
ss_sheet.title = 'Contacts'
ss.save(file_loc)
wb2.save("Rejected Contacts.xlsx")
file_loc2 = r'C:\Users\mosta\.spyder-py3\Rejected Contacts.xlsx'

ss2 = openpyxl.load_workbook(file_loc2)
ss_sheet2 = ss2.get_sheet_by_name('Sheet')
ss_sheet2.title = 'Contacts'
ss2.save(file_loc2)

